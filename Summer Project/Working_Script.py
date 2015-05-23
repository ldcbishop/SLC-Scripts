"""
This is a small script to migrate Excel files with on format to a format acceptable to the Filemaker system.

This file is being made during May of 2015 at the Behest of the Sanger Learning Center.

Primary Coder: Logan D.C. Bishop

Any questions/concerns about this code can be directed to ldcbishop@gmail.com

Working version of openpyxl requires python 2.2 usage.  Can be run with "python" command.
"""
import sys #here to allow argument passing
import os #included to splite filename from extension
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook

src_filename = sys.argv[1]
wb_origin = load_workbook(src_filename) #First argument should be script names
wb = Workbook() #Create destination workbook
dest_filename = os.path.splitext(src_filename)[0]+"_parsed.xlsx"
#Test print to see if system is pulling the correct name 
#print src_filename
#print dest_filename

