"""
This is a small script to migrate Excel files with on format to a format acceptable to the Filemaker system.

This file is being made during May of 2015 at the Behest of the Sanger Learning Center.

Primary Coder: Logan D.C. Bishop

Any questions/concerns about this code can be directed to ldcbishop@gmail.com

Working version of openpyxl requires python 2.2 usage.  Can be run with "python" command.

Not robust for widely varied excel files 
"""


import sys #here to allow argument passing
import os #included to splite filename from extension
import re #regex package
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook


"""
A student class which stores all variables in reference to a single student
"""
class Student:
	name = ""
	unique_num = 0
	uteid = ""

	def __init__(self, StuName, Stu_uteid, StuUN):
		self.name = re.sub(';', ',', StuName)
		self.unique_num = StuUN
		self.uteid = Stu_uteid
		#print "I made a student!!"

	def returnSelf(self):
		return (self.name, self.uteid, self.unique_num)

	def printMyVals(self):
		print 'My name is: '+ self.name, '\nMy UTeid is: ' + self.uteid, '\nMy unique # is: ' + str(self.unique_num)

class OutputSheet:
	mySheet = ''
	row_ptr = 1
	myCCYYS = ''

	def __init__(self, sheet, semester):
		self.mySheet = sheet
		self.myCCYYS = semester
		#Print header to output sheet
		self.mySheet['A1'] = 'ccyys'
		self.mySheet['B1'] = 'eid'
		self.mySheet['C1'] = 'name'
		self.mySheet['D1'] = 'unique'
		self.mySheet['E1'] = 'date'
		self.row_ptr = 2

	def printInfoToSheet(self, name, uteid, unique_num, attn_date):
		printRow = self.row_ptr
		self.mySheet.cell(row = printRow, column = 1).value = self.myCCYYS
		self.mySheet.cell(row = printRow, column = 2).value = uteid
		self.mySheet.cell(row = printRow, column = 3).value = name
		self.mySheet.cell(row = printRow, column = 4).value = unique_num
		self.mySheet.cell(row = printRow, column = 5).value = attn_date
		self.row_ptr = printRow + 1

	def sayName(self):
		return self.mySheet.title

"""
A function to find a key word in an excel spreadsheet and return its coordinate
@param: worksheet, and a valid string to look for.
@return: a coordinate string
"""
def searchWorkBook(work_sheet, search_string):
	#Assume Unique values are in the first 30 columns and first 50 rows
	search_space = list(work_sheet.iter_rows('A1:M50'))
	cell_re = re.compile('[A-Z]+[0-9]+')
	search_string = search_string.lower()
	for x in search_space:
		for y in x:
			if search_string in str(y.value).lower():
				return cell_re.findall(str(y))[0]
	return 'A1'

"""
A function to take the string excel coordinate and translate it into numerical coordinates
@param: cell, a coordinate string
@return: tuple with numerical coordinates (x, y)
"""
def translateCoord(cell):
	#Take a string, return a numerical tuple
	alphabetConvString = 62
	row = cell[1:]
	column = ord(cell[0]) - alphabetConvString
	return (int(row), int(column))

"""
Take a date in the origianl format e.g. 9_16 and return a properly formatted date_origin
@param: bad_date, a badly formatted date string
@return: good_date, a properly formatted date string 
"""
def parseAndBuildDate(rough_date, year):
	bad_date = str(rough_date)
	if(len(bad_date) > 10):
		return bad_date[5:7] + '/' + bad_date[8:10] + '/' +year
	alt_date_re = re.compile('[0-9]+\/[0-9]+\/[0-9]+')
	date_re = re.compile('[0-9]+\/[0-9]+')
	month_day = re.search(date_re, bad_date).group(0).split('/')
	return month_day[0]+'/'+month_day[1]+'/'+year

"""
Take a pair of coordinates for a unique student and return pertinent values
@param: r and c, which correspond to row and column
@return: a tuple of the name, uteid, and unique numbers
"""
def getStudentVals(r, c):
	name = str(sheet.cell(row = r, column = c).value)
	uteid = str(sheet.cell(row = r, column = c + 1).value)
	unique = sheet.cell(row = r, column = c + 2).value
	return (name, uteid, unique)

"""
System Main Method
"""

"""
Command format: python Working_Script.py myWork_book.xlsx
"""

src_filename = sys.argv[1]
wb_origin = load_workbook(src_filename, data_only=True) #First argument should be script names
wb_output = Workbook() #Create destination workbook
	#sheet_output = wb_output.active() #pulls active worksheet to pool data
#Work book relations are now 1:1 with each sheet being 1:1
print "Starting process for: " + src_filename
for sheet in wb_origin :

	print '\tBeginning work on sheet: ' + sheet.title
	if(sheet['D4'].value <= 0):
		print '\t\tNo students in ', src_filename, " :: ", sheet.title + '\n'
		continue
	
	student_dict = {}
	semester_year = sheet['A1'].value
	year_re = re.compile('20[0-9][0-9]')

	try:
		year = year_re.findall(semester_year)[0]
	except TypeError:
		print '\t\tNon-Attendance Sheet'
		continue
	except IndexError:
		print '\t\tMalformed sheet: ' + sheet.title
		continue

	#Two cases, either fall or Spring
	fall_re = re.compile('[Ff]+[Aa][Ll][Ll]')
	semester = fall_re.search(semester_year)
	if str(type(semester)) == "<type '_sre.SRE_Match'>" :
		CCYYS = year + '8'
	else :
		CCYYS = year + '2'


	temp_sheet = wb_output.create_sheet()
	temp_sheet.title = sheet.title + "_parsed"
	active_sheet = OutputSheet(temp_sheet, CCYYS)

	#Grabs list of all unique numbers in the spreadsheet
	un_string = sheet['A2'].value
	un_re = re.compile('[0-9][0-9][0-9][0-9][0-9]')
	unique_Num_list = un_re.findall(str(un_string))

	#Where the word date is located
	date_origin = translateCoord(searchWorkBook(sheet, 'Date'))

	#Name is leftmost cell, followed by EID and Unique #
	temp = translateCoord(searchWorkBook(sheet, 'uteid'))
	student_origin = (temp[0], temp[1] - 1)
	#print student_origin
	#All the dates loop
	current_column = date_origin[1] + 1
	loop_date = sheet.cell(row = date_origin[0], column = current_column).value

	#Loops through all dates in a sheet
	attendance_check = 0
	while(loop_date != None):
		current_row = student_origin[0] + 1
		#check to see if ANY students attended

		daily_attendance = sheet.cell(row = date_origin[0] + 1, column = current_column).value
		if(daily_attendance == 0):
			print '\t\tNo students showed up for sheet: ' + sheet.title +' :: ' + str(loop_date)
			current_column = current_column + 1
			loop_date = sheet.cell(row = date_origin[0], column = current_column).value
			continue

		if(str(loop_date).lower() == 'total session atendance'):
			total_attend = sheet.cell(row = date_origin[0] + 1, column = current_column).value 
			break
		current_date = parseAndBuildDate(loop_date,year)

		#Loops through all students in a sheet
		student_loop_check = sheet.cell(row = current_row, column = student_origin[1]).value 
		interim_attendance_check = attendance_check
		while(student_loop_check != None):

			#Check to see if the student attended
			try:
				att = int(sheet.cell(row = current_row, column = current_column).value)
			except Exception, e:
				#print 'There was a NoneType'
				#print student_loop_check
				att = 0
			
			if(att == 0):
				current_row = current_row + 1
				student_loop_check = sheet.cell(row = current_row, column = student_origin[1]).value
				#print "Current_row: " + str(current_row) + " Student value: " + str(sheet.cell(row = current_row-1, column = current_column).value)
				continue

			if(current_row not in student_dict.keys()):
				current_student = Student(*getStudentVals(current_row, student_origin[1]-2))
				student_dict[current_row] = current_student

			student_info = student_dict[current_row].returnSelf()
			active_sheet.printInfoToSheet(*student_info, attn_date = current_date)
			attendance_check = attendance_check + 1
			current_row = current_row + 1
			student_loop_check = sheet.cell(row = current_row, column = student_origin[1]).value 
			#create a write to new sheet function
		if(attendance_check - interim_attendance_check != daily_attendance):
			print '********The attendance values do not match; Daily attendence is ' + str(daily_attendance) + 'and the counted attendence is ' + str(attendance_check - interim_attendance_check) + ' on ' + current_date 
		current_column = current_column + 1
		loop_date = sheet.cell(row = date_origin[0], column = current_column).value
	print ''
#Save all changes to the workbook at the end
	if(total_attend != attendance_check):
		print '********Total Attendance: '+ str(total_attend) + ' Counted Attendance: ' + str(attendance_check)
dest_filename = os.path.splitext(src_filename)[0]+"_parsed.xlsx" #Name that the new workbook will be saved under
wb_output.remove_sheet(wb_output.active)
wb_output.save(dest_filename)
print 'Finished ' + src_filename + '\n'

"""
Current print format
ccyys, eid, name, unique, date

Getting a cell:
cell = sheet.cell(row = x, column = y)

Setting a value:
cell.value = "hello, world"

cmd string:
python Working_Script.py myWork_book.xlsx
"""



