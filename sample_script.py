from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook() #Creates an initial workbook which can be operated upon

dest_filename = 'myWork_book.xlsx' #The filename of my new workbook; String with file extension

wsl = wb.active #referenceable workbook

wsl.title = "range names" 

for row in range(1, 40):
	wsl.append(range(600))	#iterates through rows 1-40 assigning a value

ws2 = wb.create_sheet(title="Pi") #creates subsheet tab

ws2['F5'] = 3.14 #Directly references cell in a worksheet, assigning the value of 3.14

ws3 = wb.create_sheet(title = 'Data')

wb.save(filename = dest_filename)
